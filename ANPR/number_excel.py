import cv2
import datetime
import openpyxl
import pytesseract
import os

pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

platecascade = cv2.CascadeClassifier("C:\haarcascade_russian_plate_number.xml")
minArea = 500 
cap = cv2.VideoCapture(0)
count = 0

# Check if the Excel file already exists
if os.path.isfile("C:/ANPR Dataset/number.xlsx"):
    # Open the existing workbook
    workbook = openpyxl.load_workbook("C:/ANPR Dataset/number.xlsx")
else:
    # Create a new workbook if the file does not exist
    workbook = openpyxl.Workbook()

# Select the active worksheet
worksheet = workbook.active

# Find the last row with data and set the count to start from the next row
count = worksheet.max_row - 1

# Add column headers to the worksheet if it is empty
if count == 0:
    worksheet.cell(row=1, column=1).value = "ID"
    worksheet.cell(row=1, column=2).value = "Date"
    worksheet.cell(row=1, column=3).value = "Time"
    worksheet.cell(row=1, column=4).value = "Plate"

while True:
    success, img = cap.read()
    imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2BGRA)
    numberPlates = platecascade.detectMultiScale(imgGray, 1.1, 4)
    for (x, y, w, h) in numberPlates:
        area = w * h
        if area > minArea:
            cv2.rectangle(img, (x, y), (x + w, y + h), (255, 0, 0), 2)
            cv2.putText(img, "NumberPlate", (x, y - 5), cv2.FONT_HERSHEY_COMPLEX, 1, (0, 0, 255), 2)
            imgRoi = img[y:y + h, x:x + w]
            cv2.imshow("ROI", imgRoi)
            
            # Get the current date and time
            now = datetime.datetime.now()
            date_str = now.strftime("%Y-%m-%d")
            time_str = now.strftime("%H-%M-%S")

            if cv2.waitKey(1) & 0xFF == ord('s'):
                # Store the detected number plate text and the current date and time in the Excel file
                number_plate = "Number Plate " + str(count+1)
                worksheet.cell(row=count+2, column=1).value = number_plate
                worksheet.cell(row=count+2, column=2).value = date_str
                worksheet.cell(row=count+2, column=3).value = time_str
                
                # Perform OCR on the cropped image to extract the text
                text = pytesseract.image_to_string(imgRoi, lang='eng', config='--psm 11')
                worksheet.cell(row=count+2, column=4).value = text.strip()
                count += 1

                # Save the cropped number plate image and display a message on the screen
                img_path = "C:\ANPR Dataset\Plates\image - {} {} {}.jpg".format(number_plate, date_str, time_str)
                cv2.imwrite(img_path, imgRoi)
                cv2.rectangle(img, (0, 200), (640, 300), (0, 255, 0), cv2.FILLED)
                cv2.putText(img, "Scan Saved", (15, 265), cv2.FONT_HERSHEY_COMPLEX, 2, (0, 255, 255), 2)
                cv2.imshow("FINAL OUTPUT", img)
                cv2.waitKey(500)
    
    cv2.imshow("RESULT", img)
    
    key = cv2.waitKey(1) & 0xFF
    # if the `q` key was pressed, break from the loop and save the Excel file
    if key == ord('q'):
        workbook.save("C:/ANPR Dataset/number.xlsx")
        break

cv2.destroyAllWindows()
cap.release()
